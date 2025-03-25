import pandas as pd
from typing import List, Dict, Optional
import os
from datetime import datetime

class ExcelExporter:
    def export_to_excel(self, data: List[Dict], output_file: str, 
                       existing_excel: Optional[Dict] = None, append_mode: bool = False, 
                       sheet_name: str = None):
        """导出数据到Excel，保留原有格式"""
        # 创建数据框
        df = pd.DataFrame(data)

        # 追加模式：在现有Excel文件中追加数据
        if append_mode and existing_excel and os.path.exists(existing_excel['file']):
            try:
                import shutil
                # 创建原文件的备份
                backup_file = existing_excel['file'] + '.bak'
                shutil.copy2(existing_excel['file'], backup_file)

                try:
                    # 使用openpyxl打开原文件以保留所有格式
                    from openpyxl import load_workbook
                    from copy import copy
                    wb = load_workbook(existing_excel['file'])
                    
                    # 处理工作表选择
                    if sheet_name and sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                    else:
                        ws = wb.active
                    
                    # 获取用户指定的标题行
                    header_row = existing_excel['header_row'] + 1  # 从0开始转为从1开始

                    # 读取标题行的内容，确保使用正确的标题行
                    column_headers = []
                    column_indices = {}  # 存储列名与索引的映射
                    for idx, cell in enumerate(ws[header_row], 1):
                        if cell.value:
                            column_headers.append(cell.value)
                            column_indices[cell.value] = idx
                    
                    # 使用pandas读取数据（从标题行开始）
                    existing_df = pd.read_excel(
                        existing_excel['file'], 
                        header=existing_excel['header_row'],
                        sheet_name=sheet_name if sheet_name else 0,
                        engine='openpyxl'
                    )

                    # 准备新数据
                    current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    if '追加时间' in df.columns:
                        df = df.drop('追加时间', axis=1)
                    
                    # 确保新数据的列与现有数据一致
                    df = df.reindex(columns=existing_df.columns)
                    df['追加时间'] = current_time

                    # 获取最后一行的实际行号（考虑合并单元格）
                    # 从标题行开始查找，避免被前面的合并单元格干扰
                    last_row = header_row
                    actual_data_rows = 0
                    for row_idx in range(header_row + 1, ws.max_row + 1):
                        row_empty = True
                        for cell in ws[row_idx]:
                            if cell.value is not None:
                                row_empty = False
                                break
                        if not row_empty:
                            last_row = row_idx
                            actual_data_rows += 1
                    
                    # 如果没有实际数据行，使用标题行作为参考
                    data_row = header_row + 1
                    if actual_data_rows == 0:
                        # 没有数据行，使用标题行下一行
                        data_row = header_row + 1
                    else:
                        # 找到第一个有数据的行作为样式参考
                        for row_idx in range(header_row + 1, ws.max_row + 1):
                            row_has_data = False
                            for cell in ws[row_idx]:
                                if cell.value is not None:
                                    row_has_data = True
                                    break
                            if row_has_data:
                                data_row = row_idx
                                break

                    # 将新数据逐行写入
                    for idx, row in df.iterrows():
                        current_row = last_row + idx + 1
                        for col_name, col_idx in column_indices.items():
                            # 创建新单元格
                            cell = ws.cell(row=current_row, column=col_idx)
                            
                            # 获取值
                            if col_name == '追加时间':
                                value = current_time
                            else:
                                value = row.get(col_name, "")
                            cell.value = value
                            
                            # 复制样式（优先使用数据行样式，如果没有则使用标题行样式）
                            source_cell = ws.cell(row=data_row, column=col_idx)
                            if source_cell:
                                try:
                                    # 使用copy函数复制样式
                                    if hasattr(source_cell, 'has_style') and source_cell.has_style:
                                        cell._style = copy(source_cell._style)
                                except:
                                    # 如果样式复制失败，尝试使用标题行样式
                                    try:
                                        header_cell = ws.cell(row=header_row, column=col_idx)
                                        if hasattr(header_cell, 'has_style') and header_cell.has_style:
                                            cell._style = copy(header_cell._style)
                                    except:
                                        pass

                    # 检查并应用合并单元格
                    self._handle_merged_cells(ws, header_row, last_row, df)

                    # 保存工作簿
                    wb.save(output_file)
                    # 如果保存成功，删除备份文件
                    if os.path.exists(backup_file):
                        os.remove(backup_file)

                except Exception as e:
                    # 如果出错，恢复备份
                    if os.path.exists(backup_file):
                        shutil.copy2(backup_file, existing_excel['file'])
                    raise e

            except Exception as e:
                # 确保清理备份文件
                backup_file = existing_excel['file'] + '.bak'
                if os.path.exists(backup_file):
                    if not os.path.exists(existing_excel['file']):
                        shutil.copy2(backup_file, existing_excel['file'])
                    os.remove(backup_file)
                raise Exception(f"追加到Excel时出错: {str(e)}")

        # 新建模式：创建新的Excel文件
        else:
            if 'key' in df.columns and 'value' in df.columns:
                df = df[['filename', 'folder', 'key', 'value']]
            df.to_excel(output_file, index=False, engine='openpyxl')
            
    def _handle_merged_cells(self, worksheet, header_row, last_row, new_data):
        """处理合并单元格"""
        # 查找数据行的第一列是否有合并单元格，如果有则为新增数据添加相同的合并
        merged_cols = []
        for merged_range in worksheet.merged_cells.ranges:
            if merged_range.min_row > header_row and merged_range.min_col == 1:
                # 找到数据行第一列的合并单元格
                merged_cols.append((
                    merged_range.min_col, 
                    merged_range.max_col, 
                    merged_range.max_row - merged_range.min_row + 1
                ))
        
        # 为新增数据应用相同的合并规则
        for merged_col_info in merged_cols:
            min_col, max_col, height = merged_col_info
            for i in range(len(new_data)):
                start_row = last_row + i * height + 1
                end_row = start_row + height - 1
                if min_col != max_col:  # 只有当实际有多个列合并时才执行
                    worksheet.merge_cells(
                        start_row=start_row, 
                        start_column=min_col, 
                        end_row=end_row, 
                        end_column=max_col
                    )
    
    def get_excel_sheets(self, excel_file: str) -> List[str]:
        """获取Excel文件中所有工作表的名称"""
        if not os.path.exists(excel_file):
            return []
            
        try:
            # 使用openpyxl打开Excel文件
            from openpyxl import load_workbook
            wb = load_workbook(excel_file, read_only=True)
            return wb.sheetnames
        except Exception as e:
            print(f"获取工作表出错: {str(e)}")
            return []
